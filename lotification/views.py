from django.shortcuts import render,redirect
from django.http import HttpResponse
from django.contrib import messages
from .forms  import *
from django.db import connection
from .filters import *
from django.contrib.auth.forms import UserCreationForm
import django_excel as excel
from datetime import datetime
from openpyxl import Workbook
from django.http.response  import HttpResponse
from io import BytesIO
# from reportlab.lib.pagesizes import A4, cm
from reportlab.pdfgen import canvas
from django.http import FileResponse
from django.conf import settings

from django.contrib.auth.models import Group

# login
from .decorators import *
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required


# Create your views here.
@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def home(request):


    title = "Home"
    users = User_empl.objects.all()
    users_total = User_empl.objects.all().count()
    group = request.user.groups.all()[0].name
    clients = Clients.objects.raw('SELECT * FROM lotification_Clients LIMIT 5')
    clients_count = Clients.objects.all().count()

    lotes = Pots.objects.raw('SELECT * FROM lotification_Pots LIMIT 5')
    pots_count = Pots.objects.all().count()
    

    if group == 'Admin':
        permisison=True
    else:
        permisison= False
    
    
    context = {'lotes':lotes,
               'clients':clients,
               'numero_clientes':clients_count,
               'numero_lotes':pots_count,
               'permission':permisison, 
               'title':title,
               'users':users,
               'users_total':users_total,
               }   
    return render(request, 'lotification/home.html',context)



@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def user_list(request):
    title = "Usuarios"
    users = User_empl.objects.all()
    users_total = User_empl.objects.all().count()
    context = {'users':users,
               'title':title,
               'users_total':users_total,
               }
    return render(request, 'lotification/user_list.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def customers_list(request):
    title = "Clientes"
    clients = Clients.objects.all()
    
    cliente_filtrado = ClientFilter(request.GET, queryset=clients)
    clients = cliente_filtrado.qs
    print(cliente_filtrado)
    context = {'clients':clients,'filtro':cliente_filtrado}
    return render(request, 'lotification/customer_list.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def lote_list(request):
    title = "Lotes"
    lotes = Pots.objects.all()
    myFilter = PotFilter(request.GET, queryset=lotes) 

    lotes = myFilter.qs  
    context = {'lotes':lotes,'filtro':myFilter, 'title':title}
    return render(request, 'lotification/lote_list.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def lote_info(request,pk):
    title = "Lote "+ str(pk) + " info "
    cursor = Pots.objects.get(id = pk)
    form = PaymentForm(initial={'Payment_Pot':cursor})
    # pass
    # query= '''SELECT * FROM lotification_Clients WHERE id = 1'''
    # cursor = connection.cursor()
    # cursor.execute(query)
    # print(cursor)
   
    all_Pots = Pots.objects.all()
    # x = cursor.payments_set.all()
    if cursor.pot_map == None:
        print('NO HAY IMAGEN')
    else:
        print('HAY MAPA')
    pagos = Payments.objects.filter(Payment_Pot = cursor.id)
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/lote')
    area = cursor.pot_large * cursor.pot_width
    # params = {'area:'}
    abono_total = 0

    if pagos:
        for pago in pagos:
            abono_total +=pago.Payment_amount
            print(pago.Payment_amount)
    else:
        print('No hay pagos')

    print(str(abono_total))
    restante = cursor.pot_price - abono_total
    context = {'lote':cursor, 'pagos':pagos,'form_payment':form, 'area':area, 'abono':abono_total, 'restante':restante,'id_pot':pk, 'title':title}

    return render(request, 'lotification/lote_info.html',context)



@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def new_lote(request):
    form = PotForm()
    if request.method == 'POST':
        form = PotForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/lote')
    context = {'form':form}
    return render(request,'lotification/new_lote.html',context  )


@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def new_client(request):
    # pass
    form = ClientsForm()

    if request.method == 'POST':
        # print(request.POST)
        form = ClientsForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/clientes')
    context = {'form':form}
    return render(request,'lotification/new_client.html',context)

@unauthenticated_user
def login_page(request):
    title = 'Inicio de Sesion'
    if request.method =='POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request,username= username, password=password)
        if user is not None:
            login(request,user)
            return redirect('home')
        else: 
            messages.info(request, 'Username or password are incorrect')
    context = {'title':title}
    return render(request, 'lotification/login.html', context)


@allowed_users(allowed_roles = ['Admin'])
def register(request):

    form = CreateUserForm()
    if request.method == 'POST':
        form = CreateUserForm(request.POST)
        if form.is_valid():
            user = form.save()
            username = form.cleaned_data.get('username')    
            group = Group.objects.get(name="Users")
            user.groups.add(group)
            User_empl.objects.create(
                user_user_emplo=user,
            )
            messages.success(request,'Cuenta creada para '+ username)
            return redirect('login')
        else:
            for msg in form.error_messages:
                messages.error(request, f"{msg}: {form.error_messages[msg]}")
                print(msg)

    context = {'form':form}
    return render(request, 'lotification/register.html',context)


def log_out(request):
    logout(request)
    return redirect('login')


# def list(request):
#     lotes = Lote.objects.all()

#     return render(request,'lotification/list.html',{'lotes':lotes})


@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def gen_EXCEL(request, pk):
    cursor = Pots.objects.get(id = pk)
    pagos = Payments.objects.filter(Payment_Pot = cursor.id)
    wb = Workbook()
    ws = wb.active
    ws['B1'] = "Reporte pagos"
    ws.merge_cells('B1:D1')
    ws['B3'] = 'ID de Pago'
    ws['C3'] = 'Monto ($)'
    ws['D3'] = 'Fecha'
    contador = 4
    for pago in pagos:
        ws.cell(row = contador, column = 2).value = pago.id
        ws.cell(row = contador, column = 3).value = pago.Payment_amount
        ws.cell(row = contador, column = 4).value = "{0:%d-%m-%Y}".format(pago.Payment_date)
        contador+=1
    today = datetime.now()
    strToday = today.strftime("%Y%m%d")
    nombre_archivo = "Reporte_pagos_Lote"+str(cursor.id)+ strToday+".xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    content = "attachment; filename = {0}".format(nombre_archivo)
    response['Content-Disposition']=content
    wb.save(response)
    return response 

    # sheet = excel.pe.Sheet(export)
    # sheet.group_rows_by_column(0)
    # print(export)
    # return excel.make_response(sheet,"csv",file_name="results de pago-"+strToday+".csv")

@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def pdf_report(request, pk):
    lote = Pots.objects.get(id=pk)
    today = datetime.now()
    strToday = today.strftime("%d/%m/%Y")
    str_name = today.strftime("%d%m%Y")
    buffer = BytesIO()
    p = canvas.Canvas(buffer)

    #Header
    #30 like margin in the x axis is a good number for a Paragraph
    p.setLineWidth(.3)
    p.setFont('Helvetica',22)
    p.drawString(30,760,"Funeraria Jerusalen")
    
    
    # img = str(settings.BASE_DIR) + "\static\img\logo.jpg"
    # print(img)
    # p.drawImage(img,30,700,80,80)


    p.setFont('Helvetica',12)
    p.drawString(30,730,"Cotizacion")
    #Date
    p.drawString(500,760,strToday)

    p.setFont('Helvetica',22)
    Pot_id = "Lote Numero " + str(lote.id)
    p.drawString(250, 700,Pot_id)


    p.setFont('Helvetica-Bold', 12)
    p.drawString(50,640,"Largo: ")
    p.drawString(50,610,"Ancho: ")
    p.drawString(50,580,"Area: ")
    p.drawString(50,550,"Precio: ")
    p.drawString(50,520,"Plazo: ")
    p.drawString(50,490,"Comprador: ")
    p.drawString(50,460,"Vendedor: ")


    p.setFont('Helvetica', 12)
    p.drawString(130,640,str(lote.pot_large) + " m")
    p.drawString(130,610,str(lote.pot_width) + " m")
    area= lote.pot_large * lote.pot_width
    p.drawString(130,580,str(area) + "m2")
    p.drawString(130,550, "$"+ str(lote.pot_price))
    p.drawString(130,520," 36 meses")
    p.drawString(130,490,"--------")
    p.drawString(130,460,"---------")




    p.showPage()
    p.save()
    buffer.seek(0)
    
    
    nombre_archivo = "Cotizacion_"+ str_name + "lote_numero_"+ str(lote.id) + ".pdf"
    return FileResponse(buffer,as_attachment=True,filename=nombre_archivo)



@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def edit_pot(request,pk):
    Pot = Pots.objects.get(id=pk)
    form = PotForm(instance=Pot)
    context = {
        'form': form,
    }
    if request.method == 'POST':
        form = PotForm(request.POST,request.FILES ,instance=Pot)
        if form.is_valid():
            form.save()
            return redirect('/lote')

    return render(request, 'lotification/edit_pot.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def edit_payment(request,pk,pk_pay):
    Payment = Payments.objects.get(id=pk_pay)
    form = PaymentForm(instance=Payment)
    context = {
        'form': form,
    }
    if request.method == 'POST':
        form = PaymentForm(request.POST,instance=Payment)
        if form.is_valid():
            form.save()
            return redirect('/lote_info/'+str(pk))
    return render(request, 'lotification/edit_payment.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def delete_pot(request,pk):
    pot = Pots.objects.get(id = pk)
    if request.method == "POST":
        pot.delete()
        return redirect('/lote')
    return render(request, 'lotification/delete_pot.html',{'pot':pot})


@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def delete_payment(request,pk,pk_pay):
    Pay = Payments.objects.get(id = pk_pay)
    if request.method =="POST":
        Pay.delete()
        return redirect('/lote_info/'+str(pk))
    return render (request, 'lotification/delete_payment.html',{'pay':Pay})


@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def delete_client(request,pk):
    client = Clients.objects.get(id = pk)
    if request.method == "POST":
        client.delete()
        return redirect('/clientes')
    return render(request, 'lotification/delete_client.html',{'client':client})

@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def edit_client(request, pk):
    Client = Clients.objects.get(id= pk)
    form = ClientsForm(instance=Client)
    context = {
        'form':form
    }
    if request.method == 'POST':
        form = ClientsForm(request.POST,instance=Client)
        if form.is_valid():
            form.save()
            return redirect('/clientes')
    return render(request, 'lotification/edit_client.html',context)



@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin','Users'])
def edit_user(request,pk):
    Current_user = User_empl.objects.get(user_user_emplo=pk)
    # print(Current_user.user_user_emplo)
    form_u = UserEmplForm(instance=Current_user)
    # print(form_u)
    context = {
        'form': form_u,
        'current_user':Current_user,
    }
    if request.method == 'POST':
        form = UserEmplForm(request.POST ,instance=Current_user)
        if form.is_valid():
            print('VALID FORM')
            form.save()
            return redirect('home')
    return render(request, 'lotification/User_profile.html',context)


@login_required(login_url='login')
@allowed_users(allowed_roles = ['Admin'])
def delete_user(request,pk):
    user = User.objects.get(id = pk)
    if request.method == "POST":
        user.delete()
        return redirect('/user_list')
    return render(request, 'lotification/delete_user.html',{'user':user})
